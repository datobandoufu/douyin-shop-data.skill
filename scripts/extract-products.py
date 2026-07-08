#!/usr/bin/env python3
# 从交易明细 xlsx 的「商品构成」sheet 提取商品名，供品牌校验使用。
# 用法: python extract-products.py <file.xlsx>
# 输出 JSON: {"products": [...商品名...], "sample": [...全表商品词样本...]}
import sys, json
from openpyxl import load_workbook

# 注意：不能用 read_only=True。罗盘导出的 xlsx 在 read_only 模式下读不到数据
#（行索引/结构特殊），必须用普通模式，否则会误判成空模板。
fp = sys.argv[1]
wb = load_workbook(fp, data_only=True)

# 主数据源：商品构成 sheet，只提取「商品名称」列（避开金额/编号/维度等噪声）
target = next((s for s in wb.sheetnames if "商品构成" in s), None)
products = []
if target:
    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        header = rows[0]
        name_col = 1  # 默认第 2 列
        for i, h in enumerate(header):
            if h and '商品名称' in str(h):
                name_col = i
                break
        for row in rows[1:]:
            if name_col < len(row) and row[name_col] is not None:
                t = str(row[name_col]).strip()
                if len(t) >= 2:
                    products.append(t)
    products = sorted(set(products))

# 兜底：商品构成为空时，全 workbook 扫描可能含商品名的单元格（商品名常含这些词）
sample = []
KEYWORDS = ("粮", "罐", "冻干", "零食", "猫", "狗", "宠", "肉", "饲料", "饼干")
for s in wb.sheetnames:
    ws = wb[s]
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if c is None:
                continue
            t = str(c).strip()
            if any(k in t for k in KEYWORDS) and t not in sample:
                sample.append(t)
wb.close()

print(json.dumps({"products": products, "sample": sample[:200]}, ensure_ascii=False))
