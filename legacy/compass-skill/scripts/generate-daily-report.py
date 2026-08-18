#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成「日期_品牌_日报.xlsx」——抖店电商罗盘每日报表聚合。

读取某品牌某日的 3 个源文件，按规则生成 4 个 sheet：
  1. 交易明细-成交概览 : 成交概览 sheet，筛选 载体类型=全部 且 投放时段=不限
  2. 交易明细-自营成交 : 自营成交 sheet，仅筛选 投放时段=不限（不限定载体类型）
  3. 商品列表          : 商品列表全量数据
  4. 达人列表          : 达人列表，筛选 成交商品 != 0（非0成交）

用法:
  python generate-daily-report.py --date 20260707 --brand 品牌A
  python generate-daily-report.py --date 20260707 --brand 品牌B

依赖: openpyxl（pip install openpyxl；或用 --project 指定目录，PROJECT_DIR 环境变量亦可覆盖）
"""
import sys
import os
import argparse
from openpyxl import load_workbook, Workbook

# 项目根目录（源文件与产出都在 项目根/日期/ 下）；默认当前工作目录，可用环境变量 PROJECT_DIR 覆盖
PROJECT_DIR = os.environ.get("PROJECT_DIR") or os.getcwd()

# 报表 4 个 sheet 的命名
SHEET_TX_OVERVIEW = "交易明细-成交概览"
SHEET_TX_SELF = "交易明细-自营成交"
SHEET_PRODUCT = "商品列表"
SHEET_INFLUENCER = "达人列表"


def find_col(headers, name):
    """按表头名精确查找列索引（忽略首尾空白）。"""
    for i, h in enumerate(headers):
        if h is not None and str(h).strip() == name:
            return i
    return None


def read_sheet_rows(fp, sheet=None):
    """读取整张表为 [(row)...]，保留表头。若 sheet=None 取第一张表。"""
    wb = load_workbook(fp, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return list(ws.iter_rows(values_only=True))


def copy_rows(ws, rows):
    for r in rows:
        ws.append(r)


def filter_tx(rows, carrier_val="全部", period_val="不限"):
    """交易明细类表：筛选 投放时段=period_val 的行；若 carrier_val 非空，再限定 载体类型=carrier_val。

    carrier_val=None 表示不限定载体类型（仅按 投放时段 过滤）。
    """
    if not rows:
        return []
    headers = rows[0]
    ci_carrier = find_col(headers, "载体类型")
    ci_period = find_col(headers, "投放时段")
    out = [headers]
    if ci_period is None:
        print(f"  [WARN] 未找到 投放时段 列，返回空数据")
        return out
    for r in rows[1:]:
        if str(r[ci_period]).strip() != period_val:
            continue
        if carrier_val is not None:
            if ci_carrier is None or str(r[ci_carrier]).strip() != carrier_val:
                continue
        out.append(r)
    return out


def filter_influencer(rows):
    """达人列表：保留 成交商品 != 0 的行。"""
    if not rows:
        return []
    headers = rows[0]
    ci = find_col(headers, "成交商品")
    out = [headers]
    if ci is None:
        print(f"  [WARN] 未找到 成交商品 列，返回全量")
        return rows
    for r in rows[1:]:
        v = r[ci]
        try:
            num = float(v) if v is not None else 0
        except (ValueError, TypeError):
            num = 0
        if num != 0:
            out.append(r)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", required=True, help="日期 YYYYMMDD")
    ap.add_argument("--brand", required=True, help="品牌简称（与文件名一致，如 品牌A / 品牌B）")
    ap.add_argument("--project", default=PROJECT_DIR)
    args = ap.parse_args()

    date, brand = args.date, args.brand
    base = os.path.join(args.project, date)
    fp_tx = os.path.join(base, f"{date}_{brand}_交易明细.xlsx")
    fp_prod = os.path.join(base, f"{date}_{brand}_商品列表.xlsx")
    fp_infl = os.path.join(base, f"{date}_{brand}_达人列表.xlsx")

    for fp in (fp_tx, fp_prod, fp_infl):
        if not os.path.exists(fp):
            print(f"[ERROR] 缺少源文件: {fp}")
            sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 成交概览
    print(f"[1/4] {SHEET_TX_OVERVIEW} ...")
    rows = read_sheet_rows(fp_tx, "成交概览")
    out = filter_tx(rows)
    ws1 = wb.create_sheet(SHEET_TX_OVERVIEW)
    copy_rows(ws1, out)
    print(f"      匹配行数: {ws1.max_row - 1}")

    # Sheet 2: 自营成交（仅 投放时段=不限，不限定载体类型）
    print(f"[2/4] {SHEET_TX_SELF} ...")
    rows = read_sheet_rows(fp_tx, "自营成交")
    out = filter_tx(rows, carrier_val=None)
    ws2 = wb.create_sheet(SHEET_TX_SELF)
    copy_rows(ws2, out)
    print(f"      匹配行数(投放时段=不限,不限载体): {ws2.max_row - 1}")

    # Sheet 3: 商品列表（全量）
    print(f"[3/4] {SHEET_PRODUCT} ...")
    rows = read_sheet_rows(fp_prod)
    ws3 = wb.create_sheet(SHEET_PRODUCT)
    copy_rows(ws3, rows)
    print(f"      行数: {ws3.max_row - 1}")

    # Sheet 4: 达人列表（成交商品 != 0）
    print(f"[4/4] {SHEET_INFLUENCER} ...")
    rows = read_sheet_rows(fp_infl)
    out = filter_influencer(rows)
    ws4 = wb.create_sheet(SHEET_INFLUENCER)
    copy_rows(ws4, out)
    print(f"      保留行数(成交商品≠0): {ws4.max_row - 1} (原始 {len(rows) - 1})")

    out_fp = os.path.join(base, f"{date}_{brand}_日报.xlsx")
    try:
        wb.save(out_fp)
        print(f"[OK] 已生成: {out_fp}")
    except PermissionError:
        # 目标文件被 Excel 等占用（独占锁），改写到 .part 临时文件，避免覆盖失败丢数据
        tmp_fp = out_fp + ".part"
        wb.save(tmp_fp)
        print(f"[WARN] 目标文件被占用（可能 Excel 已打开该日报），已生成临时文件:")
        print(f"        {tmp_fp}")
        print(f"        请关闭 Excel 中该日报后重新运行本脚本，或手动将 .part 改名为 .xlsx。")
    print(f"      sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
